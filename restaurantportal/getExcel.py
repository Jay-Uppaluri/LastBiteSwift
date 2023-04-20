from datetime import datetime, timedelta
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
import openpyxl

def get_firestore_database():
    cred = credentials.Certificate('C:\\Users\\srina\\Desktop\\lastbite_python\\serviceAccountKey.json')
    firebase_admin.initialize_app(cred)
    return firestore.client()

def get_orders(database):
    current_time = datetime.now()
    two_weeks_ago = current_time - timedelta(days=14)
    return database.collection("orders").where("timestamp", ">=", two_weeks_ago).get()

def get_restaurant(database, restaurant_id):
    restaurant_ref = database.collection("restaurants").document(restaurant_id)
    restaurant_doc = restaurant_ref.get()
    return restaurant_doc.to_dict() if restaurant_doc.exists else None

def group_orders_by_restaurant(orders, database):
    restaurant_amounts = {}
    for order in orders:
        order_dict = order.to_dict()
        restaurant_id = order_dict["restaurantId"]
        amount = order_dict["amount"]
        restaurant = get_restaurant(database, restaurant_id)
        if restaurant:
            restaurant_name = restaurant["name"]
            if restaurant_name in restaurant_amounts:
                restaurant_amounts[restaurant_name]["orders"].append(order_dict)
                restaurant_amounts[restaurant_name]["amount"] += amount
            else:
                restaurant_amounts[restaurant_name] = {"orders": [order_dict], "amount": amount}
    return restaurant_amounts

def write_orders_to_excel(restaurant_name, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Payment Intent ID"
    ws["B1"] = "Status"
    ws["C1"] = "Timestamp"
    ws["D1"] = "User ID"
    ws["E1"] = "Amount"

    row = 2
    for order_dict in data["orders"]:
        ws.cell(row=row, column=1, value=order_dict["paymentIntentId"])
        ws.cell(row=row, column=2, value=order_dict["status"])
        ws.cell(row=row, column=3, value=order_dict["timestamp"].strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row, column=4, value=order_dict["userId"])
        ws.cell(row=row, column=5, value=order_dict["amount"])
        row += 1

    ws["G1"] = "Total Amount Owed"
    ws["G2"] = data["amount"]

    wb.save(f"{restaurant_name}_orders.xlsx")

def main():
    database = get_firestore_database()
    orders = get_orders(database)
    restaurant_amounts = group_orders_by_restaurant(orders, database)
    for restaurant_name, data in restaurant_amounts.items():
        write_orders_to_excel(restaurant_name, data)

if __name__ == '__main__':
    main()
